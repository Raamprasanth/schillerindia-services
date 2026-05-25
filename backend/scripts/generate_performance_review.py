import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def fill_rate(cell, value):
    if value is None:
        cell.value = None
    else:
        cell.value = float(value)


def fill_number(cell, value):
    cell.value = int(value or 0)


def fill_score(cell, value):
    cell.value = int(round(value or 0))


def fill_common_header(ws, payload):
    ws["E1"] = f"For the month of {payload.get('monthLabel', '')}"


def fill_activity_rows(ws, rows, start_row):
    for idx, row in enumerate(rows):
        sheet_row = start_row + idx
        fill_number(ws[f"B{sheet_row}"], row.get("total"))
        fill_number(ws[f"C{sheet_row}"], row.get("withinTarget"))
        fill_rate(ws[f"E{sheet_row}"], row.get("prevRate"))
        fill_rate(ws[f"F{sheet_row}"], row.get("nextRate"))


def fill_division_workbook(wb, payload):
    sheet_name = payload.get("sheetName", "Apr")
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        source_name = next((name for name in wb.sheetnames if name != "NOTES"), wb.sheetnames[-1])
        ws = wb.copy_worksheet(wb[source_name])
        ws.title = sheet_name

    fill_common_header(ws, payload)
    ws["E2"] = payload.get("division", "")
    fill_activity_rows(ws, payload.get("activityRows", []), 5)

    compliance = payload.get("compliance", {})
    fill_score(ws["F18"], compliance.get("weeklyCrm"))
    fill_score(ws["B19"], compliance.get("pendingActivity"))
    fill_score(ws["D19"], compliance.get("nonSaleable"))
    fill_score(ws["F23"], compliance.get("criticalPending"))
    fill_score(ws["B24"], compliance.get("supplierWarranty"))
    fill_score(ws["D24"], compliance.get("supplierPendingReview"))
    fill_score(ws["F26"], compliance.get("fiveSRate"))
    fill_score(ws["B28"], compliance.get("purchaseIndent"))
    fill_score(ws["D28"], compliance.get("quarterlyBuyback"))
    fill_score(ws["F28"], compliance.get("repairReport"))

    narratives = payload.get("narratives", {})
    ws["B32"] = narratives.get("justification", "")
    ws["B33"] = narratives.get("corrective", "")
    ws["B34"] = narratives.get("hod", "")


def fill_person_workbook(wb, payload):
    ws = wb["PR - Final"]
    fill_common_header(ws, payload)
    employee = payload.get("employee", "")
    division = payload.get("employeeDivision") or payload.get("division") or ""
    ws["E2"] = f"{employee} - {division}".strip(" -")
    fill_activity_rows(ws, payload.get("activityRows", []), 5)

    row14 = payload.get("row14")
    row15 = payload.get("row15")
    if row14:
      fill_number(ws["B14"], row14.get("total"))
      fill_number(ws["C14"], row14.get("withinTarget"))
      fill_rate(ws["E14"], row14.get("prevRate"))
      fill_rate(ws["F14"], row14.get("nextRate"))
    if row15:
      fill_number(ws["B15"], row15.get("total"))
      fill_number(ws["C15"], row15.get("withinTarget"))
      fill_rate(ws["E15"], row15.get("prevRate"))
      fill_rate(ws["F15"], row15.get("nextRate"))

    compliance = payload.get("compliance", {})
    fill_score(ws["F20"], compliance.get("weeklyCrm"))
    fill_score(ws["B21"], compliance.get("pendingActivity"))
    fill_score(ws["D21"], compliance.get("nonSaleable"))
    fill_score(ws["F25"], compliance.get("criticalPending"))
    fill_score(ws["B26"], compliance.get("supplierWarranty"))
    fill_score(ws["D26"], compliance.get("supplierPendingReview"))
    fill_score(ws["F28"], compliance.get("fiveSRate"))
    fill_score(ws["B30"], compliance.get("purchaseIndent"))
    fill_score(ws["D30"], compliance.get("quarterlyBuyback"))
    fill_score(ws["F30"], compliance.get("repairReport"))

    narratives = payload.get("narratives", {})
    ws["B38"] = narratives.get("justification", "")
    ws["B39"] = narratives.get("corrective", "")
    ws["B40"] = narratives.get("hod", "")


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: generate_performance_review.py <payload_json> <template_path> <output_path>")

    payload_path = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    wb = load_workbook(template_path, keep_vba=True)

    if payload.get("scope") == "division":
        fill_division_workbook(wb, payload)
    else:
        fill_person_workbook(wb, payload)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(str(output_path))


if __name__ == "__main__":
    main()
