import json
import os
import smtplib
import ssl
import sys
from copy import copy
from email.message import EmailMessage
from email.utils import formatdate
from pathlib import Path


HEADERS = [
    "SC_ENGINEER",
    "FRN_NO",
    "BRANCH",
    "ENGINEER_ID",
    "CUST_NAME",
    "PRODUCT_MODEL",
    "UNIT_STATUS",
    "DEF_MOD_BRD_NAME",
    "DEF_GIR_NO",
    "REP_GIR_NO",
    "DEF_UNIT_GIR_NO",
    "FINAL_REMARKS",
    "DESTINATION",
    "SHIPMENT REF NUMBER",
    "REF DATE",
]
TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "escalation_formats.xlsx"


def truncate(value, limit=42):
    text = str(value or "")
    return text if len(text) <= limit else text[: limit - 3] + "..."


def get_headers(rows, preferred=None):
    headers = list(preferred or [])
    if headers:
        return headers
    discovered = []
    for row in rows:
        for key in row.keys():
            if key not in discovered:
                discovered.append(key)
    return discovered


def build_pdf(payload, output_path):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    title = payload.get("subject", "Escalation Report")
    body_lines = [line.strip() for line in payload.get("body", "").splitlines() if line.strip()]

    story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    story.append(Spacer(1, 4 * mm))
    for line in body_lines:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 5 * mm))

    column_widths = [
        22 * mm, 20 * mm, 22 * mm, 22 * mm, 30 * mm,
        26 * mm, 20 * mm, 28 * mm, 18 * mm, 18 * mm,
        24 * mm, 28 * mm, 22 * mm, 26 * mm, 26 * mm,
    ]

    for index, sheet in enumerate(payload.get("sheets", []), start=1):
        rows = sheet.get("rows", [])
        story.append(Paragraph(f"<b>{sheet.get('name', f'Section {index}')}</b>", styles["Heading2"]))
        story.append(Spacer(1, 2 * mm))

        table_rows = [HEADERS]
        for row in rows:
            table_rows.append([truncate(row.get(header, "")) for header in HEADERS])

        if len(table_rows) == 1:
            table_rows.append(["No records"] + [""] * (len(HEADERS) - 1))

        table = Table(table_rows, colWidths=column_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#12344D")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEADING", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#8EA9C1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFD")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        story.append(Spacer(1, 6 * mm))

    doc.build(story)


def build_xlsx(payload, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    title_font = Font(bold=True, color="12344D")

    for index, sheet in enumerate(payload.get("sheets", []), start=1):
        ws = wb.create_sheet(title=(sheet.get("name", f"Sheet {index}")[:31]))
        rows = sheet.get("rows", [])
        headers = get_headers(rows, sheet.get("headers"))
        if not headers:
            headers = ["Message"]
            rows = [{"Message": "No records"}]

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in rows:
            ws.append([row.get(header, "") for header in headers])

        for column_cells in ws.columns:
            max_length = 12
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value) + 2, 40))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length

        ws.freeze_panes = "A2"

    wb.save(output_path)


def copy_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def with_wrap(alignment, default_horizontal=None, default_vertical="center"):
    from openpyxl.styles import Alignment

    return Alignment(
        horizontal=alignment.horizontal or default_horizontal,
        vertical=alignment.vertical or default_vertical,
        text_rotation=alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=alignment.shrink_to_fit,
        indent=alignment.indent,
    )


def template_headers(template_ws, header_row):
    headers = []
    for col in range(1, template_ws.max_column + 1):
        value = template_ws.cell(header_row, col).value
        if value is None and not any(template_ws.cell(row, col).value is not None for row in range(1, min(template_ws.max_row, 5) + 1)):
            continue
        headers.append(str(value or "").strip())
    while headers and not headers[-1]:
        headers.pop()
    return headers


def build_template_xlsx(payload, output_path):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    template_file = Path(payload.get("templatePath") or os.getenv("ESCALATION_TEMPLATE_PATH") or TEMPLATE_PATH)
    if not template_file.exists():
        build_xlsx(payload, output_path)
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    template_wb = load_workbook(template_file)
    wb = Workbook()
    wb.remove(wb.active)

    for index, sheet in enumerate(payload.get("sheets", []), start=1):
        template_name = sheet.get("template") or sheet.get("name")
        if template_name not in template_wb.sheetnames:
            template_name = template_wb.sheetnames[0]
        template_ws = template_wb[template_name]
        title = (sheet.get("name") or template_name or f"Sheet {index}")[:31]
        ws = wb.create_sheet(title=title)
        header_row = int(sheet.get("headerRow") or 1)
        headers = sheet.get("headers") or template_headers(template_ws, header_row)
        rows = sheet.get("rows", [])

        for col, header in enumerate(headers, start=1):
            source_cell = template_ws.cell(header_row, col)
            target_cell = ws.cell(1, col, header)
            copy_cell_style(source_cell, target_cell)
            target_cell.alignment = with_wrap(source_cell.alignment, default_horizontal="center", default_vertical="center")

            letter = target_cell.column_letter
            if template_ws.column_dimensions[letter].width:
                ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width

        if template_ws.row_dimensions[header_row].height:
            ws.row_dimensions[1].height = template_ws.row_dimensions[header_row].height

        if rows:
            for row_index, row in enumerate(rows, start=2):
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row_index, col, row.get(header, ""))
                    template_data_cell = template_ws.cell(header_row + 1, col)
                    copy_cell_style(template_data_cell, cell)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            ws.cell(2, 1, "No records")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def send_email(payload, attachment_path):
    smtp_host = os.getenv("ESCALATION_SMTP_HOST", "").strip()
    smtp_port = int(os.getenv("ESCALATION_SMTP_PORT", "587") or "587")
    smtp_user = os.getenv("ESCALATION_SMTP_USER", "").strip()
    smtp_pass = os.getenv("ESCALATION_SMTP_PASS", "").strip()
    from_addr = os.getenv("ESCALATION_EMAIL_FROM", smtp_user).strip()
    payload_to = payload.get("to", []) if isinstance(payload, dict) else []
    to_addrs = [str(x).strip() for x in payload_to if str(x).strip()]
    if not to_addrs:
        to_addrs = [x.strip() for x in os.getenv("ESCALATION_EMAIL_TO", "").split(",") if x.strip()]
    cc_addrs = [x.strip() for x in os.getenv("ESCALATION_EMAIL_CC", "").split(",") if x.strip()]
    use_ssl = os.getenv("ESCALATION_SMTP_SSL", "false").strip().lower() == "true"
    use_starttls = os.getenv("ESCALATION_SMTP_STARTTLS", "true").strip().lower() != "false"

    if not smtp_host or not from_addr or not to_addrs:
        raise RuntimeError("Email settings are incomplete. Set ESCALATION_SMTP_HOST, ESCALATION_EMAIL_FROM and ESCALATION_EMAIL_TO.")

    msg = EmailMessage()
    msg["Subject"] = payload.get("subject", "Escalation Report")
    msg["From"] = from_addr
    msg["To"] = ", ".join(to_addrs)
    if cc_addrs:
        msg["Cc"] = ", ".join(cc_addrs)
    msg["Date"] = formatdate(localtime=True)
    msg.set_content(payload.get("body", "Please find the attached escalation report."))

    with open(attachment_path, "rb") as fh:
      data = fh.read()
    subtype = "pdf" if attachment_path.suffix.lower() == ".pdf" else "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    msg.add_attachment(
        data,
        maintype="application",
        subtype=subtype,
        filename=attachment_path.name,
    )

    recipients = to_addrs + cc_addrs
    try:
        if use_ssl:
            with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ssl.create_default_context()) as server:
                if smtp_user:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg, from_addr=from_addr, to_addrs=recipients)
            return

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            if use_starttls:
                server.starttls(context=ssl.create_default_context())
                server.ehlo()
            if smtp_user:
                server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=from_addr, to_addrs=recipients)
    except smtplib.SMTPAuthenticationError as exc:
        detail = ""
        try:
            detail = exc.smtp_error.decode("utf-8", errors="ignore") if exc.smtp_error else ""
        except Exception:
            detail = str(exc)
        if "gmail.com" in smtp_host.lower():
            raise RuntimeError("Gmail SMTP authentication failed. Use a Gmail App Password instead of the normal Gmail password.") from exc
        raise RuntimeError(f"SMTP authentication failed. {detail}".strip()) from exc
    except (smtplib.SMTPConnectError, smtplib.SMTPServerDisconnected, TimeoutError, OSError) as exc:
        raise RuntimeError(f"SMTP connection failed: {exc}") from exc


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Usage: send_escalation_mail.py <input-json> <output-report>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    payload = json.loads(input_path.read_text(encoding="utf-8"))

    if str(payload.get("format", "")).lower() == "xlsx" or output_path.suffix.lower() == ".xlsx":
        if any(sheet.get("template") for sheet in payload.get("sheets", [])):
            build_template_xlsx(payload, output_path)
        else:
            build_xlsx(payload, output_path)
    else:
        build_pdf(payload, output_path)
    send_email(payload, output_path)
    print(str(output_path))


if __name__ == "__main__":
    main()
